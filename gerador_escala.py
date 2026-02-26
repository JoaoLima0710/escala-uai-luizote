#!/usr/bin/env python3
"""
Gerador Automático de Escala Médica - UAI Luizote (Psiquiatria)
Período: Dia 16 do mês atual → Dia 15 do mês seguinte

Uso: python3 gerador_escala.py <ano> <mes>
Exemplo: python3 gerador_escala.py 2026 2   → gera escala Fev/Março 2026
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date, timedelta
import calendar
import sys
import copy

# ============================================================
# CONSTANTES
# ============================================================
DIAS_SEMANA_PT = ['Seg', 'Ter', 'Qua', 'Qui', 'Sex', 'Sab', 'Dom']
MESES_PT = ['', 'Janeiro', 'Fevereiro', 'Março', 'Abril', 'Maio', 'Junho',
            'Julho', 'Agosto', 'Setembro', 'Outubro', 'Novembro', 'Dezembro']
MESES_ABREV = ['', 'Jan', 'Fev', 'Março', 'Abril', 'Maio', 'Jun',
               'Jul', 'Ago', 'Set', 'Out', 'Nov', 'Dez']

# Cores
COR_VERDE_ESCURO = 'FF3FAF46'
COR_VERDE_CLARO = 'FFA9D18E'
COR_BRANCO = 'FFFFFFFF'
COR_CINZA = 'FFB2B2B2'
COR_ROSA_LICENCA = 'FFE134FB'
COR_AMARELO_ATESTADO = 'FFFFFF00'
COR_LARANJA_VAZIO = 'FFFFA07A'  # Salmon/laranja claro para slots não preenchidos

# Fontes
FONT_BOLD = Font(bold=True, size=9, color='FF000000')
FONT_NORMAL = Font(bold=False, size=9, color='FF000000')
FONT_SHIFT = Font(bold=True, size=9, color='FF000000')
FONT_TITLE = Font(bold=True, size=11, color='FF000000')
FONT_HEADER = Font(bold=True, size=9, color='FF000000')

# Alinhamentos
ALIGN_CENTER = Alignment(horizontal='center', vertical='center')
ALIGN_LEFT = Alignment(horizontal='left', vertical='center')

# Fills
FILL_VERDE_ESCURO = PatternFill('solid', fgColor=COR_VERDE_ESCURO)
FILL_VERDE_CLARO = PatternFill('solid', fgColor=COR_VERDE_CLARO)
FILL_BRANCO = PatternFill('solid', fgColor=COR_BRANCO)
FILL_CINZA = PatternFill('solid', fgColor=COR_CINZA)
FILL_ROSA = PatternFill('solid', fgColor=COR_ROSA_LICENCA)
FILL_AMARELO = PatternFill('solid', fgColor=COR_AMARELO_ATESTADO)
FILL_LARANJA = PatternFill('solid', fgColor=COR_LARANJA_VAZIO)

# Bordas
THIN_BORDER = Border(
    left=Side(style='thin', color='FF999999'),
    right=Side(style='thin', color='FF999999'),
    top=Side(style='thin', color='FF999999'),
    bottom=Side(style='thin', color='FF999999')
)

# ============================================================
# CONFIGURAÇÃO DOS MÉDICOS
# ============================================================
MEDICOS_CONFIG = [
    {
        'nome': 'Gustavo Garcia Gonçalves',
        'matricula': '9300414',
        'meta': 8,
        'regra': 'gustavo',
    },
    {
        'nome': 'Mariana Zanatta Bechara',
        'matricula': '9300762',
        'meta': 8,
        'regra': 'mariana',
    },
    {
        'nome': 'Maurício Rosa de Almeida Junior',
        'matricula': '9300788',
        'meta': 4,
        'regra': 'mauricio',
    },
    {
        'nome': 'Sergio Monteiro Faim',
        'matricula': '9300962',
        'meta': 4,
        'regra': 'faim',
    },
    {
        'nome': 'Melissa Maria R Nascimento',
        'matricula': '9300794',
        'meta': 8,
        'regra': 'melissa',
    },
    {
        'nome': 'Bruna Silva Freitas',
        'matricula': '9301856',
        'meta': None,
        'regra': 'bruna',
    },
    {
        'nome': 'Laura Jorge Diniz Povoa',
        'matricula': '9100505',
        'meta': 0,
        'regra': 'licenca',
    },
    {
        'nome': 'Valquiria Alves Souza',
        'matricula': '8403489',
        'meta': 14.5,
        'regra': 'valquiria',
    },
]

RPA_NOMES = [
    {'nome': 'Marina Anovazzi Silva', 'matricula': 'RPA'},
    {'nome': 'Lucas Valadares Motta', 'matricula': 'RPA'},
    {'nome': 'Eduardo', 'matricula': 'RPA'},
    {'nome': 'Mayara Gois', 'matricula': 'RPA'},
    {'nome': 'Patricia', 'matricula': 'RPA'},
]


# ============================================================
# FUNÇÕES AUXILIARES
# ============================================================
def get_periodo(ano, mes):
    """Retorna lista de datas do período 16/mes até 15/mes+1."""
    inicio = date(ano, mes, 16)
    if mes == 12:
        fim = date(ano + 1, 1, 15)
    else:
        fim = date(ano, mes + 1, 15)
    dias = []
    d = inicio
    while d <= fim:
        dias.append(d)
        d += timedelta(days=1)
    return dias


def ultimo_dia_semana_no_mes(ano, mes, weekday):
    """Retorna a data do último dia com dado weekday (0=Seg) no mês/ano."""
    ultimo_dia = calendar.monthrange(ano, mes)[1]
    d = date(ano, mes, ultimo_dia)
    while d.weekday() != weekday:
        d -= timedelta(days=1)
    return d


def primeiro_dia_semana_no_mes(ano, mes, weekday):
    """Retorna a data do primeiro dia com dado weekday no mês/ano."""
    d = date(ano, mes, 1)
    while d.weekday() != weekday:
        d += timedelta(days=1)
    return d


def filtrar_por_weekday(dias, weekday):
    """Filtra lista de datas por dia da semana (0=Seg, 1=Ter, ... 6=Dom)."""
    return [d for d in dias if d.weekday() == weekday]


def meses_no_periodo(dias):
    """Retorna lista ordenada de tuplas (ano, mes) presentes no período."""
    return sorted(set((d.year, d.month) for d in dias))


def contar_plantoes(escala_medico):
    """Conta plantões em 12h. D=1, N=1, M=0.5, T=0.5, D/N=2, T/N=1.5, V=2."""
    valores = {
        'N': 1, 'D': 1, 'M': 0.5, 'T': 0.5,
        'D/N': 2, 'T/N': 1.5, 'V': 2
    }
    total = 0
    for turno in escala_medico.values():
        if turno and turno in valores:
            total += valores[turno]
    return total


def contar_plantoes_unidade(escala_medico):
    """Conta quantos 'slots de plantão' (cada 12h = 1 plantão)."""
    return contar_plantoes(escala_medico)


def nome_periodo(ano, mes):
    """Gera string do período, ex: 'Fev/Março 2026'."""
    if mes == 12:
        return f'{MESES_ABREV[mes]}/{MESES_PT[1]} {ano + 1}'
    return f'{MESES_ABREV[mes]}/{MESES_PT[mes + 1]} {ano}'


# ============================================================
# REGRAS DE DISTRIBUIÇÃO
# ============================================================

def regra_gustavo(dias):
    """
    Gustavo: 8 plantões
    - Toda segunda à noite (N)
    - Prefere sábados 24h (D/N) no 4º final de semana
    - Tentar 2º e 4º final de semana: domingos D para completar
    """
    escala = {}
    meta = 8
    plantoes = 0

    # 1. Segundas à noite (N)
    for d in filtrar_por_weekday(dias, 0):  # 0 = Segunda
        escala[d] = 'N'
        plantoes += 1

    sabados = filtrar_por_weekday(dias, 5)
    domingos = filtrar_por_weekday(dias, 6)

    # 2. 4º sábado D/N (24h = 2 plantões de 12h)
    if len(sabados) > 3 and plantoes + 2 <= meta:
        escala[sabados[3]] = 'D/N'
        plantoes += 2

    # 3. Completar com domingos D do 2º e 4º FDS
    dom_pref = []
    if len(domingos) > 1:
        dom_pref.append(domingos[1])  # 2º domingo
    if len(domingos) > 3:
        dom_pref.append(domingos[3])  # 4º domingo

    for d in dom_pref:
        if plantoes >= meta:
            break
        if d not in escala:
            escala[d] = 'D'
            plantoes += 1

    # 4. Se ainda faltar, completar com outros domingos D
    for d in domingos:
        if plantoes >= meta:
            break
        if d not in escala:
            escala[d] = 'D'
            plantoes += 1

    return escala, plantoes


def regra_mariana(dias, ativa=True):
    """
    Mariana: 8 plantões
    - Toda quinta à noite (N)
    - Todo domingo à noite (N) até completar 8
    - Começar contagem do primeiro domingo
    """
    escala = {}
    if not ativa:
        return escala, 0

    meta = 8
    plantoes = 0

    # Quintas à noite
    for d in filtrar_por_weekday(dias, 3):  # 3 = Quinta
        if plantoes < meta:
            escala[d] = 'N'
            plantoes += 1

    # Domingos à noite para completar
    for d in filtrar_por_weekday(dias, 6):  # 6 = Domingo
        if plantoes >= meta:
            break
        if d not in escala:
            escala[d] = 'N'
            plantoes += 1

    return escala, plantoes


def regra_mauricio(dias):
    """
    Maurício: 4 plantões
    - Todas as sextas do período (N), EXCETO a última sexta de cada mês calendário
    - Primeiro sábado do mês novo no período (D = 12h dia)
    - Quando tem 5 sextas, fica 2 sextas sem plantão
    """
    escala = {}
    meta = 4
    plantoes = 0

    meses = meses_no_periodo(dias)

    # Identificar sextas proibidas (última sexta de cada mês)
    sextas_proibidas = set()
    for (a, m) in meses:
        ultima_sex = ultimo_dia_semana_no_mes(a, m, 4)  # 4 = Sexta
        if ultima_sex in dias:
            sextas_proibidas.add(ultima_sex)

    # Verificar se há 5 sextas (precisa pular 2)
    sextas_periodo = filtrar_por_weekday(dias, 4)
    sextas_validas = [s for s in sextas_periodo if s not in sextas_proibidas]

    # Se 5 sextas no total e já tiramos as últimas de cada mês,
    # verificar se ainda temos mais de 3 sextas válidas
    # (precisamos de no máximo 3 sextas N + 1 sábado D = 4)
    if len(sextas_validas) > 3:
        # Manter apenas as 3 primeiras sextas válidas
        sextas_validas = sextas_validas[:3]

    # Atribuir N nas sextas válidas
    for sex in sextas_validas:
        if plantoes < meta:
            escala[sex] = 'N'
            plantoes += 1

    # Primeiro sábado do segundo mês (mês novo) → D (12h dia)
    if len(meses) > 1:
        (a2, m2) = meses[1]
        primeiro_sab = primeiro_dia_semana_no_mes(a2, m2, 5)  # 5 = Sábado
        if primeiro_sab in dias and plantoes < meta:
            escala[primeiro_sab] = 'D'
            plantoes += 1

    return escala, plantoes


def regra_faim(dias, escala_mauricio):
    """
    Faim: 4 plantões
    - 3 quartas à noite (N), nunca a última quarta de cada mês calendário
    - 1 sábado noite (N) - primeiro sábado do mês novo (casado com Maurício)
    """
    escala = {}
    meta = 4
    plantoes = 0

    meses = meses_no_periodo(dias)

    # Quartas proibidas (última quarta de cada mês)
    quartas_proibidas = set()
    for (a, m) in meses:
        ultima_qua = ultimo_dia_semana_no_mes(a, m, 2)  # 2 = Quarta
        if ultima_qua in dias:
            quartas_proibidas.add(ultima_qua)

    # 3 quartas (excluindo proibidas)
    quartas_validas = [q for q in filtrar_por_weekday(dias, 2) if q not in quartas_proibidas]
    for qua in quartas_validas[:3]:
        escala[qua] = 'N'
        plantoes += 1

    # Primeiro sábado do mês novo → N (noite, casado com Maurício que faz D)
    if len(meses) > 1:
        (a2, m2) = meses[1]
        primeiro_sab = primeiro_dia_semana_no_mes(a2, m2, 5)
        if primeiro_sab in dias and plantoes < meta:
            escala[primeiro_sab] = 'N'
            plantoes += 1

    return escala, plantoes


def turno_noite_ocupado(d, *escalas):
    """Verifica se alguma escala já tem turno noturno naquele dia."""
    turnos_noite = ('N', 'D/N', 'T/N', 'V')
    for esc in escalas:
        turno = esc.get(d)
        if turno in turnos_noite:
            return True
    return False


def regra_melissa(dias, escala_mauricio, escala_faim, escala_mariana, escala_gustavo, escala_valquiria=None):
    """
    Melissa: 8 plantões, evitar finais de semana
    - Terças à noite (N)
    - Quartas que Faim não cobre (N)
    - Sextas que Maurício não cobre (N)
    - Completar com domingos à noite (N) onde turno noturno estiver vago
    - Quintas como último recurso
    """
    escala = {}
    meta = 8
    plantoes = 0

    candidatos = []

    # 1) Terças à noite (prioridade alta)
    for d in filtrar_por_weekday(dias, 1):  # 1 = Terça
        candidatos.append((0, d))

    # 2) Quartas NÃO cobertas por Faim (prioridade alta)
    for d in filtrar_por_weekday(dias, 2):  # 2 = Quarta
        if d not in escala_faim:
            candidatos.append((0, d))

    # 3) Sextas NÃO cobertas por Maurício (prioridade alta)
    for d in filtrar_por_weekday(dias, 4):  # 4 = Sexta
        if d not in escala_mauricio:
            candidatos.append((0, d))

    # 4) Domingos à noite: OK se ninguém tem turno NOTURNO naquele dia
    #    (Gustavo pode ter D = dia, que não conflita com N = noite)
    for d in filtrar_por_weekday(dias, 6):  # 6 = Domingo
        if not turno_noite_ocupado(d, escala_mariana, escala_gustavo):
            candidatos.append((1, d))

    # 5) Quintas sem turno noturno ocupado (penúltimo recurso)
    for d in filtrar_por_weekday(dias, 3):  # 3 = Quinta
        if not turno_noite_ocupado(d, escala_mariana):
            candidatos.append((2, d))

    # 6) Sábados à noite como último recurso (regra diz "evitar" FDS, não proibir)
    esc_val = escala_valquiria or {}
    for d in filtrar_por_weekday(dias, 5):  # 5 = Sábado
        if not turno_noite_ocupado(d, escala_mauricio, escala_faim, escala_gustavo, esc_val):
            candidatos.append((3, d))

    # Ordenar por prioridade e depois por data
    candidatos.sort(key=lambda x: (x[0], x[1]))
    for (prio, d) in candidatos:
        if plantoes >= meta:
            break
        if d not in escala:
            escala[d] = 'N'
            plantoes += 1

    return escala, plantoes


def regra_bruna(dias):
    """
    Bruna: M (manhã 07h-13h) em todos os dias úteis (Seg-Sex).
    Sábados de plantão especial quando necessário.
    """
    escala = {}
    for d in dias:
        if d.weekday() < 5:  # Seg a Sex
            escala[d] = 'M'
    return escala, len([d for d in dias if d.weekday() < 5])


def regra_valquiria(dias, escala_mariana, escala_gustavo=None, escala_mauricio=None, escala_faim=None):
    """
    Valquiria: EXATAMENTE ≤ 14,5 plantões (não exceder).
    Ordem de preenchimento conforme prioridade:
      1. TARDES (T): todos os dias úteis Seg-Sex (0,5 plantão cada)
      2. NOTURNOS: upgrade T → T/N nas quintas sem cobertura noturna (+1,0 plantão)
      3. FINS DE SEMANA: D/N, D ou T até completar 14,5 sem duplicata
    """
    META = 14.5
    escala = {}
    esc_gus = escala_gustavo or {}
    esc_mau = escala_mauricio or {}
    esc_fai = escala_faim or {}
    total = 0.0

    # ── 1. TARDES: T em todos os dias úteis (Seg a Sex), respeitando o teto ──
    for d in sorted(d for d in dias if d.weekday() < 5):
        if total + 0.5 > META:
            break
        escala[d] = 'T'
        total += 0.5

    # ── 2. NOTURNOS: upgrade T → T/N nas quintas onde noite está disponível ──
    #   Incremento por quinta = +1,0 (N = 12 h = 1 plantão além do T já contado)
    for d in filtrar_por_weekday(dias, 3):
        if d not in escala:
            continue  # Quinta não foi alocada no passo 1 (teto já atingido)
        if turno_noite_ocupado(d, escala_mariana):
            continue  # Mariana (ou outro) já cobre a noite → mantém só T
        if total + 1.0 > META:
            break  # Adicionar noite ultrapassaria 14,5 → para
        escala[d] = 'T/N'
        total += 1.0  # Apenas o incremento noturno (T→T/N)

    # ── 3. FINS DE SEMANA: preencher o restante até 14,5 sem conflito ──
    fds = sorted(filtrar_por_weekday(dias, 5) + filtrar_por_weekday(dias, 6))
    for d in fds:
        if total >= META:
            break
        if d in escala:
            continue  # Já alocado
        if turno_noite_ocupado(d, esc_gus, esc_mau, esc_fai):
            continue  # Outro médico tem turno noturno nesse dia → pular
        restante = round(META - total, 1)
        if restante >= 2:
            escala[d] = 'D/N'
            total += 2
        elif restante >= 1:
            escala[d] = 'D'
            total += 1
        elif restante >= 0.5:
            escala[d] = 'T'
            total += 0.5

    return escala, round(total, 1)


# ============================================================
# GERADOR PRINCIPAL
# ============================================================

def gerar_escala(ano, mes, mariana_ativa=False):
    """
    Gera a escala para o período 16/mes/ano → 15/mes+1/ano.
    Retorna dict com todas as escalas e metadados.
    """
    dias = get_periodo(ano, mes)
    resultado = {}
    alertas = []

    # 1. Bruna (manhã - não conflita com turnos N)
    esc_bruna, cnt_bruna = regra_bruna(dias)
    resultado['Bruna Silva Freitas'] = esc_bruna

    # 2. Gustavo (Segunda N + fins de semana)
    esc_gustavo, cnt_gustavo = regra_gustavo(dias)
    resultado['Gustavo Garcia Gonçalves'] = esc_gustavo
    if cnt_gustavo < 8:
        alertas.append(f"Gustavo: faltam {8 - cnt_gustavo} plantão(ões)")

    # 3. Mariana (Quinta N + Domingo N)
    esc_mariana, cnt_mariana = regra_mariana(dias, ativa=mariana_ativa)
    resultado['Mariana Zanatta Bechara'] = esc_mariana
    if mariana_ativa and cnt_mariana < 8:
        alertas.append(f"Mariana: faltam {8 - cnt_mariana} plantão(ões)")

    # 4. Maurício (Sextas N + 1 Sábado D)
    esc_mauricio, cnt_mauricio = regra_mauricio(dias)
    resultado['Maurício Rosa de Almeida Junior'] = esc_mauricio
    if cnt_mauricio < 4:
        alertas.append(f"Maurício: faltam {4 - cnt_mauricio} plantão(ões)")

    # 5. Faim (3 Quartas N + 1 Sábado N)
    esc_faim, cnt_faim = regra_faim(dias, esc_mauricio)
    resultado['Sergio Monteiro Faim'] = esc_faim
    if cnt_faim < 4:
        alertas.append(f"Faim: faltam {4 - cnt_faim} plantão(ões)")

    # 6. Laura (licença)
    resultado['Laura Jorge Diniz Povoa'] = {}

    # 7. Pré-calcular as quintas de Valquiria (N se Mariana não cobre)
    #    para que Melissa saiba quais noites estão livres
    esc_valquiria_parcial = {}
    for d in filtrar_por_weekday(dias, 3):  # Quintas
        if d not in esc_mariana:
            esc_valquiria_parcial[d] = 'T/N'

    # 8. Melissa (Terças N + Wed/Fri gaps + Domingo + Sábado fallback)
    esc_melissa, cnt_melissa = regra_melissa(
        dias, esc_mauricio, esc_faim, esc_mariana, esc_gustavo, esc_valquiria_parcial)
    resultado['Melissa Maria R Nascimento'] = esc_melissa
    if cnt_melissa < 8:
        alertas.append(f"Melissa: faltam {8 - cnt_melissa} plantão(ões)")

    # 9. Valquiria (ÚLTIMA) — respeita limite 14,5 e evita conflitos com outros
    esc_valquiria, cnt_valquiria = regra_valquiria(
        dias, esc_mariana,
        escala_gustavo=esc_gustavo,
        escala_mauricio=esc_mauricio,
        escala_faim=esc_faim,
    )
    resultado['Valquiria Alves Souza'] = esc_valquiria

    # RPAs ficam vazios (preenchimento manual)
    for rpa in RPA_NOMES:
        resultado[rpa['nome']] = {}

    # Identificar slots vagos (MANHÃ, TARDE, NOITE)
    slots_vagos = {'M': {}, 'T': {}, 'N': {}}
    for d in dias:
        m_coberto = False
        t_coberto = False
        n_coberto = False

        for nome, esc in resultado.items():
            turno = esc.get(d)
            if turno in ('M', 'D', 'D/N', 'V'):
                m_coberto = True
            if turno in ('T', 'D', 'D/N', 'V', 'T/N'):
                t_coberto = True
            if turno in ('N', 'D/N', 'V', 'T/N'):
                n_coberto = True

        if not m_coberto:
            slots_vagos['M'][d] = True
        if not t_coberto:
            slots_vagos['T'][d] = True
        if not n_coberto:
            slots_vagos['N'][d] = True

    return {
        'dias': dias,
        'escalas': resultado,
        'alertas': alertas,
        'slots_vagos': slots_vagos,
        'ano': ano,
        'mes': mes,
        'mariana_ativa': mariana_ativa,
    }


# ============================================================
# GERAÇÃO DO EXCEL
# ============================================================

def gerar_excel(dados, caminho_saida):
    """Gera o arquivo Excel com a escala formatada."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'PSIQUIATRIA'

    dias = dados['dias']
    escalas = dados['escalas']
    alertas = dados['alertas']
    slots_vagos = dados['slots_vagos']
    ano = dados['ano']
    mes = dados['mes']
    num_dias = len(dias)

    # Configurar largura de colunas
    ws.column_dimensions['A'].width = 30.5
    ws.column_dimensions['B'].width = 14.9
    for i in range(num_dias):
        col_letter = get_column_letter(3 + i)  # C = 3
        ws.column_dimensions[col_letter].width = 4.6
    col_total = get_column_letter(3 + num_dias + 2)  # 2 colunas de espaço
    ws.column_dimensions[col_total].width = 9.0

    # ---- CABEÇALHO (Linhas 1-9) ----
    ws.merge_cells(start_row=1, start_column=1, end_row=9, end_column=2)
    ws.merge_cells(start_row=1, start_column=3, end_row=2, end_column=12)
    c = ws.cell(row=1, column=3, value='UAI LUIZOTE')
    c.font = FONT_TITLE
    c.alignment = ALIGN_CENTER
    c.fill = FILL_BRANCO

    ws.cell(row=3, column=3, value='PSIQUIATRIA').font = FONT_BOLD
    ws['C3'].alignment = ALIGN_CENTER
    ws['C3'].fill = FILL_BRANCO

    periodo_str = nome_periodo(ano, mes)
    ws.cell(row=5, column=3, value=periodo_str).font = FONT_BOLD
    ws['C5'].alignment = ALIGN_CENTER
    ws['C5'].fill = FILL_BRANCO

    # Legenda de horários (coluna M-T)
    col_leg = 13  # M
    legenda = [
        (2, 'V', '07:00 às 07:00', '24 Horas'),
        (3, 'D', '07:00 às 19:00', '12 Horas'),
        (4, 'N', '19:00 às 07:00', '12 Horas'),
        (5, 'M', '07:00 às 13:00', '6 Horas'),
        (6, 'T', '13:00 às 19:00', '6 Horas'),
    ]
    ws.merge_cells(start_row=1, start_column=col_leg, end_row=1, end_column=col_leg + 7)
    ws.cell(row=1, column=col_leg, value='HORÁRIOS').font = FONT_BOLD
    ws.cell(row=1, column=col_leg).alignment = ALIGN_CENTER
    ws.cell(row=1, column=col_leg).fill = FILL_BRANCO

    for (row, cod, horario, duracao) in legenda:
        ws.cell(row=row, column=col_leg, value=cod).font = FONT_BOLD
        ws.cell(row=row, column=col_leg).alignment = ALIGN_CENTER
        ws.cell(row=row, column=col_leg).fill = FILL_BRANCO
        ws.merge_cells(start_row=row, start_column=col_leg + 1, end_row=row, end_column=col_leg + 5)
        ws.cell(row=row, column=col_leg + 1, value=horario).font = FONT_BOLD
        ws.cell(row=row, column=col_leg + 1).alignment = ALIGN_CENTER
        ws.cell(row=row, column=col_leg + 1).fill = FILL_BRANCO
        ws.merge_cells(start_row=row, start_column=col_leg + 6, end_row=row, end_column=col_leg + 7)
        ws.cell(row=row, column=col_leg + 6, value=duracao).font = FONT_BOLD
        ws.cell(row=row, column=col_leg + 6).alignment = ALIGN_CENTER
        ws.cell(row=row, column=col_leg + 6).fill = FILL_BRANCO

    # SD, F, A
    ws.cell(row=7, column=col_leg, value='SD').font = FONT_BOLD
    ws.cell(row=7, column=col_leg).alignment = ALIGN_CENTER
    ws.cell(row=7, column=col_leg).fill = FILL_BRANCO
    ws.merge_cells(start_row=7, start_column=col_leg + 2, end_row=7, end_column=col_leg + 4)
    ws.cell(row=7, column=col_leg + 2, value='Folga Sindicato').font = FONT_BOLD
    ws.cell(row=7, column=col_leg + 2).alignment = ALIGN_CENTER
    ws.cell(row=7, column=col_leg + 2).fill = FILL_BRANCO

    ws.cell(row=8, column=col_leg, value='F').font = FONT_BOLD
    ws.cell(row=8, column=col_leg).alignment = ALIGN_CENTER
    ws.cell(row=8, column=col_leg).fill = FILL_BRANCO
    ws.merge_cells(start_row=8, start_column=col_leg + 1, end_row=8, end_column=col_leg + 7)
    ws.cell(row=8, column=col_leg + 1, value='Férias').font = FONT_BOLD
    ws.cell(row=8, column=col_leg + 1).alignment = ALIGN_CENTER
    ws.cell(row=8, column=col_leg + 1).fill = FILL_BRANCO

    ws.cell(row=9, column=col_leg, value='A').font = FONT_BOLD
    ws.cell(row=9, column=col_leg).alignment = ALIGN_CENTER
    ws.cell(row=9, column=col_leg).fill = FILL_BRANCO
    ws.merge_cells(start_row=9, start_column=col_leg + 1, end_row=9, end_column=col_leg + 7)
    ws.cell(row=9, column=col_leg + 1, value='Atestado/Congresso').font = FONT_BOLD
    ws.cell(row=9, column=col_leg + 1).alignment = ALIGN_CENTER
    ws.cell(row=9, column=col_leg + 1).fill = FILL_BRANCO

    # ---- LINHA 10: Médicos / MATRÍCULA ----
    ws.row_dimensions[10].height = 7.5
    ws.merge_cells(start_row=10, start_column=1, end_row=12, end_column=1)
    ws.cell(row=10, column=1, value='Médicos:').font = FONT_BOLD
    ws.cell(row=10, column=1).alignment = ALIGN_CENTER
    ws.cell(row=10, column=1).fill = FILL_BRANCO
    ws.cell(row=10, column=2, value='MATRÍCULA').font = FONT_BOLD
    ws.cell(row=10, column=2).alignment = ALIGN_CENTER
    ws.cell(row=10, column=2).fill = FILL_BRANCO

    # ---- LINHA 11: Dias da semana ----
    ws.row_dimensions[11].height = 13.5
    for i, d in enumerate(dias):
        col = 3 + i
        ws.cell(row=11, column=col, value=DIAS_SEMANA_PT[d.weekday()])
        ws.cell(row=11, column=col).font = FONT_BOLD
        ws.cell(row=11, column=col).alignment = ALIGN_CENTER
        ws.cell(row=11, column=col).fill = FILL_VERDE_ESCURO
        ws.cell(row=11, column=col).border = THIN_BORDER

    # ---- LINHA 12: Números dos dias ----
    ws.row_dimensions[12].height = 13.5
    for i, d in enumerate(dias):
        col = 3 + i
        ws.cell(row=12, column=col, value=d.day)
        ws.cell(row=12, column=col).font = FONT_BOLD
        ws.cell(row=12, column=col).alignment = ALIGN_CENTER
        ws.cell(row=12, column=col).fill = FILL_VERDE_CLARO
        ws.cell(row=12, column=col).border = THIN_BORDER

    # ---- LINHAS DE MÉDICOS (13 em diante) ----
    col_total_idx = 3 + num_dias + 2  # Coluna do total

    row_atual = 13
    medicos_ordem = [
        'Gustavo Garcia Gonçalves',
        'Mariana Zanatta Bechara',
        'Maurício Rosa de Almeida Junior',
        'Sergio Monteiro Faim',
        'Melissa Maria R Nascimento',
        'Bruna Silva Freitas',
        'Laura Jorge Diniz Povoa',
        'Valquiria Alves Souza',
    ]

    config_por_nome = {c['nome']: c for c in MEDICOS_CONFIG}

    for nome in medicos_ordem:
        ws.row_dimensions[row_atual].height = 11.25
        config = config_por_nome.get(nome, {})
        esc = escalas.get(nome, {})

        # Nome
        c = ws.cell(row=row_atual, column=1, value=nome)
        c.font = FONT_NORMAL
        c.alignment = ALIGN_LEFT

        # Aplicar cor de fundo para licença
        if config.get('regra') == 'licenca':
            c.fill = FILL_ROSA
            ws.cell(row=row_atual, column=2).fill = FILL_ROSA
            ws.cell(row=row_atual, column=col_total_idx).fill = FILL_ROSA
        elif not dados['mariana_ativa'] and nome == 'Mariana Zanatta Bechara':
            c.fill = FILL_AMARELO
            ws.cell(row=row_atual, column=2).fill = FILL_AMARELO
        else:
            c.fill = FILL_BRANCO

        # Matrícula
        mat = config.get('matricula', '')
        ws.cell(row=row_atual, column=2, value=mat)
        ws.cell(row=row_atual, column=2).font = FONT_NORMAL
        ws.cell(row=row_atual, column=2).alignment = ALIGN_CENTER

        # Turnos
        for i, d in enumerate(dias):
            col = 3 + i
            turno = esc.get(d)
            cell = ws.cell(row=row_atual, column=col)
            cell.border = THIN_BORDER
            cell.alignment = ALIGN_CENTER

            if turno:
                cell.value = turno
                cell.font = FONT_SHIFT

            # Fundo: fim de semana = verde claro
            if d.weekday() >= 5:  # Sab ou Dom
                cell.fill = FILL_VERDE_CLARO
            elif config.get('regra') == 'licenca':
                cell.fill = FILL_ROSA
            elif not dados['mariana_ativa'] and nome == 'Mariana Zanatta Bechara':
                cell.fill = FILL_AMARELO

        # Total de plantões
        meta = config.get('meta')
        total = contar_plantoes(esc)
        if meta is not None and meta > 0:
            ws.cell(row=row_atual, column=col_total_idx, value=int(total) if total == int(total) else total)
        elif config.get('regra') == 'bruna':
            ws.cell(row=row_atual, column=col_total_idx, value='1fd')
        elif config.get('regra') == 'licenca':
            ws.cell(row=row_atual, column=col_total_idx, value='1fd')

        ws.cell(row=row_atual, column=col_total_idx).font = FONT_BOLD
        ws.cell(row=row_atual, column=col_total_idx).alignment = ALIGN_CENTER
        if config.get('regra') == 'licenca':
            ws.cell(row=row_atual, column=col_total_idx).fill = FILL_ROSA

        row_atual += 1

    # ---- LINHAS RPA (vazias para preenchimento manual) ----
    for rpa in RPA_NOMES:
        ws.row_dimensions[row_atual].height = 11.25
        ws.cell(row=row_atual, column=1, value=rpa['nome']).font = FONT_NORMAL
        ws.cell(row=row_atual, column=1).fill = FILL_BRANCO
        ws.cell(row=row_atual, column=2, value='RPA').font = FONT_BOLD
        ws.cell(row=row_atual, column=2).alignment = ALIGN_CENTER
        ws.cell(row=row_atual, column=2).fill = FILL_BRANCO

        for i, d in enumerate(dias):
            col = 3 + i
            cell = ws.cell(row=row_atual, column=col)
            cell.border = THIN_BORDER
            cell.alignment = ALIGN_CENTER
            if d.weekday() >= 5:
                cell.fill = FILL_VERDE_CLARO

        row_atual += 1

    # ---- LINHA VAZIA ----
    row_atual += 1

    # ---- ATESTADO ----
    ws.merge_cells(start_row=row_atual, start_column=1, end_row=row_atual, end_column=2)
    ws.cell(row=row_atual, column=1, value='Atestado').font = FONT_BOLD
    ws.cell(row=row_atual, column=1).fill = FILL_CINZA
    ws.cell(row=row_atual, column=1).alignment = ALIGN_LEFT
    row_atual += 2

    # ---- FALTAM (alertas por médico) ----
    ws.merge_cells(start_row=row_atual, start_column=1, end_row=row_atual, end_column=2)
    ws.cell(row=row_atual, column=1, value='FALTAM').font = FONT_BOLD
    ws.cell(row=row_atual, column=1).alignment = ALIGN_CENTER
    ws.cell(row=row_atual, column=1).fill = FILL_BRANCO
    row_atual += 1

    # Mostrar alertas de plantões faltantes
    for alerta in alertas:
        ws.cell(row=row_atual, column=1, value=alerta).font = Font(bold=False, size=8, color='FFFF0000')
        ws.cell(row=row_atual, column=1).alignment = ALIGN_LEFT
        row_atual += 1

    row_atual += 1

    # ---- FALTAM (slots vagos com cor laranja) ----
    ws.merge_cells(start_row=row_atual, start_column=1,
                   end_row=row_atual, end_column=3 + num_dias - 1)
    ws.cell(row=row_atual, column=1, value='FALTAM PREENCHER (alocar RPA)').font = FONT_BOLD
    ws.cell(row=row_atual, column=1).alignment = ALIGN_CENTER
    ws.cell(row=row_atual, column=1).fill = FILL_LARANJA
    row_atual += 1

    # MANHÃ
    ws.merge_cells(start_row=row_atual, start_column=1, end_row=row_atual, end_column=2)
    ws.cell(row=row_atual, column=1, value='MANHÃ').font = FONT_BOLD
    ws.cell(row=row_atual, column=1).alignment = ALIGN_CENTER
    ws.cell(row=row_atual, column=1).fill = FILL_BRANCO
    for i, d in enumerate(dias):
        col = 3 + i
        cell = ws.cell(row=row_atual, column=col)
        cell.font = FONT_BOLD
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER
        if d in slots_vagos['M']:
            cell.value = 1
            cell.fill = FILL_LARANJA
        else:
            cell.value = 0
            cell.fill = FILL_BRANCO
    row_atual += 1

    # TARDE
    ws.merge_cells(start_row=row_atual, start_column=1, end_row=row_atual, end_column=2)
    ws.cell(row=row_atual, column=1, value='TARDE').font = FONT_BOLD
    ws.cell(row=row_atual, column=1).alignment = ALIGN_CENTER
    ws.cell(row=row_atual, column=1).fill = FILL_BRANCO
    for i, d in enumerate(dias):
        col = 3 + i
        cell = ws.cell(row=row_atual, column=col)
        cell.font = FONT_BOLD
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER
        if d in slots_vagos['T']:
            cell.value = 1
            cell.fill = FILL_LARANJA
        else:
            cell.value = 0
            cell.fill = FILL_BRANCO
    row_atual += 1

    # NOITE
    ws.merge_cells(start_row=row_atual, start_column=1, end_row=row_atual, end_column=2)
    ws.cell(row=row_atual, column=1, value='NOITE').font = FONT_BOLD
    ws.cell(row=row_atual, column=1).alignment = ALIGN_CENTER
    ws.cell(row=row_atual, column=1).fill = FILL_BRANCO
    for i, d in enumerate(dias):
        col = 3 + i
        cell = ws.cell(row=row_atual, column=col)
        cell.font = FONT_BOLD
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER
        if d in slots_vagos['N']:
            cell.value = 1
            cell.fill = FILL_LARANJA
        else:
            cell.value = 0
            cell.fill = FILL_BRANCO
    row_atual += 1

    # ---- LEGENDA ----
    row_atual += 1
    ws.cell(row=row_atual, column=1, value='LEGENDA DE CORES:').font = FONT_BOLD
    row_atual += 1
    ws.cell(row=row_atual, column=1, value='Laranja = Slots vagos para RPA').font = FONT_NORMAL
    ws.cell(row=row_atual, column=2).fill = FILL_LARANJA
    row_atual += 1
    ws.cell(row=row_atual, column=1, value='Verde = Final de semana').font = FONT_NORMAL
    ws.cell(row=row_atual, column=2).fill = FILL_VERDE_CLARO
    row_atual += 1
    ws.cell(row=row_atual, column=1, value='Rosa = Licença maternidade').font = FONT_NORMAL
    ws.cell(row=row_atual, column=2).fill = FILL_ROSA
    row_atual += 1
    ws.cell(row=row_atual, column=1, value='Amarelo = Atestado').font = FONT_NORMAL
    ws.cell(row=row_atual, column=2).fill = FILL_AMARELO

    # ---- BLOCO "DATAS PARA PREENCHER" (canto superior direito) ----
    # Posicionado 2 colunas após o total, alinhado ao topo da tabela
    col_rpa_bloco = col_total_idx + 2       # coluna inicial do bloco
    dias_semana_full = ['Segunda', 'Terça', 'Quarta', 'Quinta',
                        'Sexta', 'Sábado', 'Domingo']

    # Consolidar datas com pelo menos 1 turno vago
    todas_datas_rpa = sorted(
        set(slots_vagos['M']) | set(slots_vagos['T']) | set(slots_vagos['N'])
    )

    # Larguras das colunas do bloco
    ws.column_dimensions[get_column_letter(col_rpa_bloco)].width     = 11   # data
    ws.column_dimensions[get_column_letter(col_rpa_bloco + 1)].width = 10   # dia semana
    ws.column_dimensions[get_column_letter(col_rpa_bloco + 2)].width = 24   # turnos vagos

    FILL_LARANJA_TITULO = PatternFill('solid', fgColor='FFE65100')           # laranja escuro
    FILL_LARANJA_LINHA  = PatternFill('solid', fgColor='FFFFF3E0')           # laranja muito claro
    FILL_FDS_RPA        = PatternFill('solid', fgColor='FFFFE0B2')           # laranja/amarelo FDS
    FONT_RPA_TITULO     = Font(bold=True,  size=9, color='FFFFFFFF')
    FONT_RPA_CAB        = Font(bold=True,  size=8, color='FF7F3000')
    FONT_RPA_DATA       = Font(bold=False, size=8, color='FF212121')
    FONT_RPA_TURNO      = Font(bold=True,  size=8, color='FFE65100')
    THIN_BORDER_RPA = Border(
        left=Side(style='thin', color='FFFFCC80'),
        right=Side(style='thin', color='FFFFCC80'),
        top=Side(style='thin', color='FFFFCC80'),
        bottom=Side(style='thin', color='FFFFCC80'),
    )

    # Linha 10 — título do bloco (mesclado nas 3 colunas)
    ws.merge_cells(start_row=10, start_column=col_rpa_bloco,
                   end_row=10,   end_column=col_rpa_bloco + 2)
    c_titulo = ws.cell(row=10, column=col_rpa_bloco,
                       value='DATAS PARA PREENCHER (RPA)')
    c_titulo.font      = FONT_RPA_TITULO
    c_titulo.fill      = FILL_LARANJA_TITULO
    c_titulo.alignment = ALIGN_CENTER
    c_titulo.border    = THIN_BORDER_RPA

    # Linha 11 — sub-cabeçalho de colunas
    for col_off, label in enumerate(['Data', 'Dia', 'Turnos vagos']):
        c = ws.cell(row=11, column=col_rpa_bloco + col_off, value=label)
        c.font      = FONT_RPA_CAB
        c.fill      = PatternFill('solid', fgColor='FFFFCC80')
        c.alignment = ALIGN_CENTER
        c.border    = THIN_BORDER_RPA

    # Linha 12 — vazia (alinha com a linha de números de dias)
    for col_off in range(3):
        c = ws.cell(row=12, column=col_rpa_bloco + col_off)
        c.fill   = PatternFill('solid', fgColor='FFFFCC80')
        c.border = THIN_BORDER_RPA

    # Linhas de dados — uma por data descoberta, a partir da linha 13
    for idx, d in enumerate(todas_datas_rpa):
        row_d = 13 + idx
        fds   = d.weekday() >= 5
        fill_linha = FILL_FDS_RPA if fds else FILL_LARANJA_LINHA

        # Coluna 1: data (dd/mm)
        c_data = ws.cell(row=row_d, column=col_rpa_bloco,
                         value=d.strftime('%d/%m/%Y'))
        c_data.font      = FONT_RPA_DATA
        c_data.fill      = fill_linha
        c_data.alignment = ALIGN_CENTER
        c_data.border    = THIN_BORDER_RPA

        # Coluna 2: dia da semana
        c_dia = ws.cell(row=row_d, column=col_rpa_bloco + 1,
                        value=dias_semana_full[d.weekday()])
        c_dia.font      = FONT_RPA_DATA
        c_dia.fill      = fill_linha
        c_dia.alignment = ALIGN_CENTER
        c_dia.border    = THIN_BORDER_RPA

        # Coluna 3: turnos vagos
        vagos = []
        if d in slots_vagos['M']: vagos.append('M')
        if d in slots_vagos['T']: vagos.append('T')
        if d in slots_vagos['N']: vagos.append('N')
        c_turno = ws.cell(row=row_d, column=col_rpa_bloco + 2,
                          value='  ·  '.join(vagos))
        c_turno.font      = FONT_RPA_TURNO
        c_turno.fill      = fill_linha
        c_turno.alignment = ALIGN_CENTER
        c_turno.border    = THIN_BORDER_RPA

    # Rodapé do bloco: total de datas
    row_rod = 13 + len(todas_datas_rpa)
    ws.merge_cells(start_row=row_rod, start_column=col_rpa_bloco,
                   end_row=row_rod,   end_column=col_rpa_bloco + 2)
    total_slots = (len(slots_vagos['M']) + len(slots_vagos['T']) + len(slots_vagos['N']))
    c_rod = ws.cell(row=row_rod, column=col_rpa_bloco,
                    value=f"{len(todas_datas_rpa)} datas  |  {total_slots} slots vagos")
    c_rod.font      = FONT_RPA_CAB
    c_rod.fill      = PatternFill('solid', fgColor='FFFFCC80')
    c_rod.alignment = ALIGN_CENTER
    c_rod.border    = THIN_BORDER_RPA

    wb.save(caminho_saida)
    return caminho_saida


# ============================================================
# CORES ANSI (terminal)
# ============================================================
ANSI_RESET   = '\033[0m'
ANSI_BOLD    = '\033[1m'
ANSI_GREEN   = '\033[92m'   # Verde brilhante → meta atingida
ANSI_YELLOW  = '\033[93m'   # Amarelo → incompleto / atestado
ANSI_RED     = '\033[91m'   # Vermelho → falta crítica
ANSI_CYAN    = '\033[96m'   # Ciano → cabeçalho / seção
ANSI_GRAY    = '\033[90m'   # Cinza → licença / info
ANSI_ORANGE  = '\033[33m'   # Laranja → slot vago RPA
ANSI_WHITE   = '\033[97m'   # Branco brilhante


def _cor(texto, codigo):
    return f"{codigo}{texto}{ANSI_RESET}"


# ============================================================
# EXIBIÇÃO: CONTADOR DE PLANTÕES
# ============================================================
def exibir_contador_plantoes(dados):
    """
    Exibe tabela de contagem de plantões no terminal.
    Linhas ficam VERDES ao atingir a meta individual.
    """
    escalas = dados['escalas']
    periodo = nome_periodo(dados['ano'], dados['mes'])
    config_por_nome = {c['nome']: c for c in MEDICOS_CONFIG}

    ordem = [
        'Gustavo Garcia Gonçalves',
        'Mariana Zanatta Bechara',
        'Maurício Rosa de Almeida Junior',
        'Sergio Monteiro Faim',
        'Melissa Maria R Nascimento',
        'Bruna Silva Freitas',
        'Laura Jorge Diniz Povoa',
        'Valquiria Alves Souza',
    ]

    larg_nome  = 36
    larg_real  = 10
    larg_meta  = 6
    larg_bar   = 20
    linha_sep  = '─' * (larg_nome + larg_real + larg_meta + larg_bar + 12)

    print()
    print(_cor('═' * (larg_nome + larg_real + larg_meta + larg_bar + 12), ANSI_CYAN))
    print(_cor(f"  CONTADOR DE PLANTÕES  —  {periodo}", ANSI_BOLD + ANSI_CYAN))
    print(_cor('═' * (larg_nome + larg_real + larg_meta + larg_bar + 12), ANSI_CYAN))

    cabecalho = (
        f"  {'Médico':<{larg_nome}}"
        f"{'Realizado':>{larg_real}}"
        f"{'Meta':>{larg_meta}}"
        f"  {'Progresso':<{larg_bar}}"
        f"  Status"
    )
    print(_cor(cabecalho, ANSI_BOLD + ANSI_WHITE))
    print(_cor(linha_sep, ANSI_GRAY))

    for nome in ordem:
        esc     = escalas.get(nome, {})
        total   = contar_plantoes(esc)
        cfg     = config_por_nome.get(nome, {})
        meta    = cfg.get('meta')
        regra   = cfg.get('regra', '')

        # ── Determinar status e cor ──
        if regra == 'licenca':
            status_txt = '─  Licença/Maternidade'
            cor_linha  = ANSI_GRAY
            barra_txt  = ''
            meta_txt   = '─'
            real_txt   = f"{total:>5.1f}"

        elif not dados['mariana_ativa'] and nome == 'Mariana Zanatta Bechara':
            status_txt = '⚠  Atestado (desativada)'
            cor_linha  = ANSI_YELLOW
            barra_txt  = ''
            meta_txt   = str(int(meta)) if meta == int(meta) else str(meta)
            real_txt   = f"{total:>5.1f}"

        elif regra == 'bruna':
            status_txt = '─  Plantão fixo matutino'
            cor_linha  = ANSI_GRAY
            barra_txt  = ''
            meta_txt   = '─'
            real_txt   = f"{total:>5.1f}"

        elif meta is None:
            status_txt = '─'
            cor_linha  = ANSI_GRAY
            barra_txt  = ''
            meta_txt   = '─'
            real_txt   = f"{total:>5.1f}"

        else:
            meta_num  = float(meta)
            pct       = min(total / meta_num, 1.0) if meta_num > 0 else 1.0
            barras    = int(pct * larg_bar)
            barra_txt = '█' * barras + '░' * (larg_bar - barras)
            meta_txt  = str(int(meta)) if meta == int(meta) else str(meta)
            real_txt  = f"{total:>5.1f}"

            if total >= meta_num:
                status_txt = '✅ META ATINGIDA'
                cor_linha  = ANSI_GREEN
            elif total >= meta_num * 0.75:
                status_txt = f'⚠  Faltam {round(meta_num - total, 1)}'
                cor_linha  = ANSI_YELLOW
            else:
                status_txt = f'✗  Faltam {round(meta_num - total, 1)}'
                cor_linha  = ANSI_RED

        linha = (
            f"  {nome:<{larg_nome}}"
            f"{real_txt:>{larg_real}}"
            f"{meta_txt:>{larg_meta}}"
            f"  {barra_txt:<{larg_bar}}"
            f"  {status_txt}"
        )
        print(_cor(linha, cor_linha))

    print(_cor(linha_sep, ANSI_GRAY))


# ============================================================
# EXIBIÇÃO: RELATÓRIO DE DATAS DESCOBERTAS (RPA)
# ============================================================
def exibir_relatorio_rpa(dados):
    """
    Exibe relatório de datas com slots vagos, por turno,
    indicando o dia da semana — para oferta de RPA.
    """
    slots_vagos = dados['slots_vagos']
    periodo     = nome_periodo(dados['ano'], dados['mes'])

    # Consolidar todas as datas com pelo menos 1 vago
    todas_datas = sorted(
        set(slots_vagos['M']) | set(slots_vagos['T']) | set(slots_vagos['N'])
    )

    print()
    print(_cor('═' * 62, ANSI_ORANGE))
    print(_cor(f"  RELATÓRIO DE DATAS DESCOBERTAS  —  {periodo}", ANSI_BOLD + ANSI_ORANGE))
    print(_cor('  Turnos para oferta de RPA', ANSI_ORANGE))
    print(_cor('═' * 62, ANSI_ORANGE))

    if not todas_datas:
        print(_cor('  ✅  Todos os turnos estão cobertos!', ANSI_GREEN))
        print()
        return

    # Cabeçalho da tabela
    cab = f"  {'Data':<12}{'Dia da Semana':<16}{'M':^6}{'T':^6}{'N':^6}  Turnos vagos"
    print(_cor(cab, ANSI_BOLD + ANSI_WHITE))
    print(_cor('  ' + '─' * 58, ANSI_GRAY))

    total_m = total_t = total_n = 0

    for d in todas_datas:
        data_str    = d.strftime('%d/%m/%Y')
        dia_semana  = ['Segunda', 'Terça', 'Quarta', 'Quinta',
                       'Sexta', 'Sábado', 'Domingo'][d.weekday()]
        vago_m = d in slots_vagos['M']
        vago_t = d in slots_vagos['T']
        vago_n = d in slots_vagos['N']

        if vago_m: total_m += 1
        if vago_t: total_t += 1
        if vago_n: total_n += 1

        turno_m = _cor(' ○ ', ANSI_ORANGE) if vago_m else _cor(' ● ', ANSI_GREEN)
        turno_t = _cor(' ○ ', ANSI_ORANGE) if vago_t else _cor(' ● ', ANSI_GREEN)
        turno_n = _cor(' ○ ', ANSI_ORANGE) if vago_n else _cor(' ● ', ANSI_GREEN)

        vagos_desc = []
        if vago_m: vagos_desc.append('Manhã')
        if vago_t: vagos_desc.append('Tarde')
        if vago_n: vagos_desc.append('Noite')
        desc_str = ', '.join(vagos_desc)

        # Cor da linha: FDS em amarelo, dias úteis em branco
        cor_data = ANSI_YELLOW if d.weekday() >= 5 else ANSI_WHITE

        linha_base = f"  {data_str:<12}{dia_semana:<16}"
        print(_cor(linha_base, cor_data), end='')
        print(turno_m + turno_t + turno_n, end='')
        print(_cor(f"  {desc_str}", ANSI_ORANGE))

    print(_cor('  ' + '─' * 58, ANSI_GRAY))

    # Totais por turno
    tot_linha = (
        f"  {'TOTAL VAGOS':<28}"
        f"{total_m:^6}"
        f"{total_t:^6}"
        f"{total_n:^6}"
        f"  ({total_m + total_t + total_n} slots)"
    )
    print(_cor(tot_linha, ANSI_BOLD + ANSI_ORANGE))

    # Legenda
    print()
    print(_cor('  Legenda:', ANSI_GRAY))
    print(_cor('  ○ = Turno vago (oferecer RPA)', ANSI_ORANGE))
    print(_cor('  ● = Turno coberto', ANSI_GREEN))
    print(_cor('  M = Manhã  |  T = Tarde  |  N = Noite', ANSI_GRAY))
    print(_cor('  Linhas em amarelo = Finais de semana', ANSI_YELLOW))
    print()

    # ── Bloco copiável ───────────────────────────────────────
    dias_semana_full = ['Segunda', 'Terça', 'Quarta', 'Quinta',
                        'Sexta', 'Sábado', 'Domingo']
    sep = '─' * 48
    print(sep)
    print(f"  DATAS PARA PREENCHER — {periodo}")
    print(sep)
    for d in todas_datas:
        vagos = []
        if d in slots_vagos['M']: vagos.append('Manhã')
        if d in slots_vagos['T']: vagos.append('Tarde')
        if d in slots_vagos['N']: vagos.append('Noite')
        dia_str  = dias_semana_full[d.weekday()]
        data_str = d.strftime('%d/%m/%Y')
        turnos_str = ' · '.join(vagos)
        print(f"  {data_str}  ({dia_str})  —  {turnos_str}")
    print(sep)
    print()


# ============================================================
# EXIBIÇÃO: SEQUÊNCIA CRONOLÓGICA DE PLANTÕES (copiável)
# ============================================================
def exibir_sequencia_plantoes(dados):
    """
    Exibe, em texto puro sem cores ANSI, a sequência cronológica
    de todos os plantões de cada médico — formatada para copiar e colar.
    """
    escalas        = dados['escalas']
    periodo        = nome_periodo(dados['ano'], dados['mes'])
    config_por_nome = {c['nome']: c for c in MEDICOS_CONFIG}

    ORDEM = [
        'Gustavo Garcia Gonçalves',
        'Mariana Zanatta Bechara',
        'Maurício Rosa de Almeida Junior',
        'Sergio Monteiro Faim',
        'Melissa Maria R Nascimento',
        'Bruna Silva Freitas',
        'Laura Jorge Diniz Povoa',
        'Valquiria Alves Souza',
    ]

    separador = '=' * 62
    linha_med  = '-' * 62

    print()
    print(separador)
    print(f"  SEQUÊNCIA DE PLANTÕES — {periodo}")
    print(f"  (texto copiável — ordenado cronologicamente por médico)")
    print(separador)

    for nome in ORDEM:
        esc    = escalas.get(nome, {})
        cfg    = config_por_nome.get(nome, {})
        meta   = cfg.get('meta')
        regra  = cfg.get('regra', '')
        total  = contar_plantoes(esc)
        mat    = cfg.get('matricula', '')

        print()
        # Cabeçalho do médico
        if meta is not None and meta > 0:
            meta_fmt  = int(meta) if meta == int(meta) else meta
            total_fmt = int(total) if total == int(total) else total
            print(f"  {nome}  [{mat}]  —  {total_fmt}/{meta_fmt} plantões")
        elif regra == 'bruna':
            total_fmt = int(total) if total == int(total) else total
            print(f"  {nome}  [{mat}]  —  {total_fmt} plantões (matutino fixo)")
        elif regra == 'licenca':
            print(f"  {nome}  [{mat}]  —  Licença / Maternidade")
        else:
            print(f"  {nome}  [{mat}]")

        print(f"  {linha_med}")

        if not esc:
            if not dados['mariana_ativa'] and nome == 'Mariana Zanatta Bechara':
                print(f"  (sem plantões — atestado)")
            elif regra == 'licenca':
                print(f"  (sem plantões — licença maternidade)")
            else:
                print(f"  (sem plantões alocados)")
            continue

        # Turnos ordenados por data
        turnos = sorted(esc.items())
        for d, turno in turnos:
            dia_semana = DIAS_SEMANA_PT[d.weekday()]
            data_str   = d.strftime('%d/%m/%Y')
            # Calcular carga horária descritiva
            carga = {
                'N': '19h–07h (12h noite)',
                'D': '07h–19h (12h dia)',
                'M': '07h–13h (6h manhã)',
                'T': '13h–19h (6h tarde)',
                'D/N': '07h–07h (24h)',
                'T/N': '13h–07h (18h)',
                'V':   '07h–07h (24h)',
            }.get(turno, turno)
            print(f"  {data_str}  {dia_semana:<5}  {turno:<5}  {carga}")

    print()
    print(separador)
    print()


# ============================================================
# MAIN
# ============================================================
def main():
    if len(sys.argv) < 3:
        print("Uso: python3 gerador_escala.py <ano> <mes> [mariana_ativa=1]")
        print("Exemplo: python3 gerador_escala.py 2026 2")
        print("         (gera escala Fev/Março 2026, 16/Fev → 15/Mar)")
        sys.exit(1)

    ano = int(sys.argv[1])
    mes = int(sys.argv[2])
    # Padrão: Mariana DESATIVADA (atestado). Passe '1' para ativar.
    mariana_ativa = False
    if len(sys.argv) > 3:
        mariana_ativa = sys.argv[3] == '1'

    print(f"\nGerando escala — {nome_periodo(ano, mes)}")
    print(f"  Período : 16/{mes:02d}/{ano} → 15/{(mes % 12) + 1:02d}/{ano if mes < 12 else ano + 1}")
    print(f"  Mariana : {'Ativa' if mariana_ativa else 'Atestado (desativada)'}")

    dados = gerar_escala(ano, mes, mariana_ativa=mariana_ativa)

    # ── 1. Contador de plantões ──────────────────────────────
    exibir_contador_plantoes(dados)

    # ── 2. Relatório de datas descobertas (RPA) ──────────────
    exibir_relatorio_rpa(dados)

    # ── 3. Sequência cronológica copiável ────────────────────
    exibir_sequencia_plantoes(dados)

    # ── 4. Alertas técnicos (se houver) ─────────────────────
    if dados['alertas']:
        print(_cor('⚠  ALERTAS:', ANSI_BOLD + ANSI_RED))
        for a in dados['alertas']:
            print(_cor(f"   • {a}", ANSI_RED))
        print()

    # ── 4. Gerar Excel ───────────────────────────────────────
    if mes == 12:
        nome_arquivo = f"ESCALA UAI {MESES_ABREV[mes]}_{MESES_PT[1]} {ano + 1}.xlsx"
    else:
        nome_arquivo = f"ESCALA UAI {MESES_ABREV[mes]}_{MESES_PT[mes + 1]} {ano}.xlsx"

    caminho = f"/sessions/fervent-awesome-bardeen/mnt/Documents/{nome_arquivo}"
    gerar_excel(dados, caminho)
    print(_cor(f"  Arquivo Excel: {nome_arquivo}", ANSI_BOLD + ANSI_GREEN))
    print()
    return caminho


if __name__ == '__main__':
    main()
